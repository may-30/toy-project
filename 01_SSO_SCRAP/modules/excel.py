from openpyxl import Workbook
from datetime import datetime

def excel(sso_lists):
    # 1. initial setting.
    wb = Workbook()
    ws = wb.active
    
    # 2. set sheet title.
    ws.title = 'SSO'

    # 3. first row title.
    ws.append(['No', 'Group', 'Name'])

    # 4. input data.
    number = 0
    for idx, value in enumerate(sso_lists, start = 2):
        sso_group_name = value.split(':', 1)[0]
        sso_user_name = value.split(':', 1)[1]
        ws.cell(row = idx, column = 1).value = number + 1
        ws.cell(row = idx, column = 2).value = sso_group_name
        ws.cell(row = idx, column = 3).value = sso_user_name
        number += 1
    
    # 5. merge cell
    merge_cell_B_lists = []
    for cell in ws['B']:
        merge_cell_B_lists.append(cell.value)

    merge_count = 0
    start_cell = 1
    for row in range(1, len(merge_cell_B_lists)):
        if merge_cell_B_lists[row - 1] == merge_cell_B_lists[row]:
            merge_count += 1
        else:
            if merge_count > 0:
                ws.merge_cells(start_row = start_cell, end_row = start_cell + merge_count, start_column = 1, end_column = 1)
                ws.merge_cells(start_row = start_cell, end_row = start_cell + merge_count, start_column = 2, end_column = 2)
            merge_count = 0
            start_cell = row + 1

    if merge_count > 0:
        ws.merge_cells(start_row = start_cell, end_row = start_cell + merge_count, start_column = 1, end_column = 1)
        ws.merge_cells(start_row = start_cell, end_row = start_cell + merge_count, start_column = 2, end_column = 2)

    count = 1
    for cell in ws['A']:
        if cell.value != 'No' and cell.value != None:
            cell.value = count
            count += 1

    # 6. file name
    file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    # 7. file export
    wb.save(f'./{file_name}.xlsx')